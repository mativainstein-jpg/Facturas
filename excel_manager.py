import re
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font
from config import COLS, NUM_COLS, HEADERS_FACTURAS, EXCEL_FACTURAS, EXCEL_PROVEEDORES

_FILL_ROJO   = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
_FONT_BLANCO = Font(color='FFFFFF', bold=True)


class ExcelManager:
    """Keeps the workbook open for the duration of a processing run."""

    def __init__(self):
        try:
            if EXCEL_FACTURAS.exists():
                self.wb = openpyxl.load_workbook(str(EXCEL_FACTURAS))
            else:
                self.wb = openpyxl.Workbook()
                # Quitar la hoja por defecto para que FACTURAS se cree con
                # create_sheet (y reciba sus encabezados) igual que las demás.
                self.wb.remove(self.wb.active)
        except PermissionError:
            raise PermissionError(
                f'No se puede abrir "{EXCEL_FACTURAS.name}" porque está abierto '
                'en Excel (u otro programa). Cerralo e intentá de nuevo.'
            )

        self._ensure('FACTURAS',   HEADERS_FACTURAS)
        self._ensure('DUPLICADAS', ['Fecha', 'Nombre adjunto', 'Clave comprobante', 'Motivo', 'Thread ID'])
        self._ensure('ERRORES',    ['Fecha hora', 'Nombre adjunto', 'Nivel', 'Mensaje'])

    # ------------------------------------------------------------------
    # Setup
    # ------------------------------------------------------------------

    def _ensure(self, nombre, headers=None):
        if nombre not in self.wb.sheetnames:
            ws = self.wb.create_sheet(nombre)
            if headers:
                ws.append(headers)

    # ------------------------------------------------------------------
    # Lectura de índices
    # ------------------------------------------------------------------

    def cargar_indice_duplicados(self):
        nombres, claves = set(), set()
        ws = self.wb['FACTURAS']
        col_nombre = COLS['NOMBRE_ADJUNTO']
        col_clave  = COLS['CLAVE_COMPROBANTE']
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) >= col_nombre:
                v = row[col_nombre - 1]
                if v:
                    nombres.add(str(v))
            if len(row) >= col_clave:
                v = row[col_clave - 1]
                if v:
                    claves.add(str(v))
        return {'nombres': nombres, 'claves': claves}

    # ------------------------------------------------------------------
    # Escritura
    # ------------------------------------------------------------------

    def escribir_factura(self, datos):
        fila, cols_verificar = _armar_fila_verificada(datos)
        ws = self.wb['FACTURAS']
        ws.append(fila)

        fila_num = ws.max_row
        for col in cols_verificar:
            cell = ws.cell(row=fila_num, column=col)
            cell.fill = _FILL_ROJO
            cell.font = _FONT_BLANCO

        return cols_verificar

    def registrar_duplicado(self, nombre, clave, motivo, thread_id):
        self.wb['DUPLICADAS'].append([datetime.now(), nombre, clave, motivo, thread_id])

    def registrar_error(self, nombre, mensaje):
        self.wb['ERRORES'].append([datetime.now(), nombre, 'ERROR', mensaje])

    # ------------------------------------------------------------------
    # Persistencia
    # ------------------------------------------------------------------

    def guardar(self):
        try:
            self.wb.save(str(EXCEL_FACTURAS))
        except PermissionError:
            raise PermissionError(
                f'No se puede guardar "{EXCEL_FACTURAS.name}" porque está abierto '
                'en Excel (u otro programa). Cerralo y volvé a procesar las facturas.'
            )

    def cerrar(self):
        self.wb.close()


# ------------------------------------------------------------------
# Lectura de proveedores (independiente del workbook principal)
# ------------------------------------------------------------------

def cargar_indice_proveedores():
    """
    Lee 'proveedores.xlsx' (formato Proveedores_para_cruce) y devuelve:
      { cuit_sin_guiones: {gasto, desc_gasto, rubro, desc_rubro} }

    Estructura del Excel:
      Col A: CUIT  B: codgasto  C: Gastos  D: codrubro
      Col E: idx_rubro (tabla auxiliar)  F: Descripción Rubro
    """
    indice = {}
    if not EXCEL_PROVEEDORES.exists():
        return indice

    try:
        wb = openpyxl.load_workbook(str(EXCEL_PROVEEDORES), read_only=True)
    except PermissionError:
        raise PermissionError(
            f'No se puede leer "{EXCEL_PROVEEDORES.name}" porque está abierto '
            'en Excel (u otro programa). Cerralo e intentá de nuevo.'
        )
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    # Tabla de descripciones de rubro: col E (índice 4) → col F (índice 5)
    rubros_desc = {}
    for row in rows:
        if len(row) >= 6 and row[4] is not None and row[5] is not None:
            try:
                rubros_desc[int(row[4])] = str(row[5]).strip()
            except (ValueError, TypeError):
                pass

    # Índice de proveedores: col A → cols B-D
    for row in rows:
        if not row[0]:
            continue
        cuit = re.sub(r'\D', '', str(row[0]))
        if not cuit:
            continue
        try:
            gasto = int(row[1]) if row[1] is not None else None
        except (ValueError, TypeError):
            gasto = None
        desc_gasto = str(row[2]).strip() if row[2] else None
        try:
            rubro = int(row[3]) if row[3] is not None else None
        except (ValueError, TypeError):
            rubro = None
        desc_rubro = rubros_desc.get(rubro) if rubro is not None else None
        indice[cuit] = {
            'gasto':      gasto,
            'desc_gasto': desc_gasto,
            'rubro':      rubro,
            'desc_rubro': desc_rubro,
        }

    wb.close()
    return indice


# ------------------------------------------------------------------
# Construcción de fila
# ------------------------------------------------------------------

def _armar_fila_verificada(d):
    ahora = datetime.now()
    fila  = [None] * NUM_COLS
    cols_verificar = []

    def set_col(key, valor):
        fila[COLS[key] - 1] = valor

    set_col('TIPO_COMPROBANTE',     d['tipo'])
    set_col('NUMERO',               d['numero'])
    set_col('PUNTO_VENTA',          d['punto_venta'])
    set_col('FECHA',                d['fecha'])
    set_col('DENOMINACION',         d['denominacion'])
    set_col('CUIT',                 d['cuit'])
    set_col('NETO',                 d['neto_num'])
    set_col('IVA',                  d['iva_num'])
    set_col('NO_GRAVADO',           d.get('no_gravado_num'))
    set_col('IMP_INTERNOS',         d.get('imp_internos_num'))
    set_col('EXENTOS',              d.get('exentos_num'))
    set_col('PERCEPCION_IVA',       d.get('percepcion_iva_num'))
    set_col('PERCEPCION_IIBB',      d.get('percepcion_iibb_num'))
    set_col('KILOS',                d['kilos_num'])
    set_col('PRECIO_UNITARIO',      d['precio_unitario_num'])
    set_col('MONOTRIBUTISTA',       d['monotributista'])
    set_col('PERCEPCION_GANANCIAS', d.get('percepcion_ganancias_num'))
    set_col('TOTAL',                d['total_num'])
    set_col('TASA',                 d['tasa'])
    set_col('GASTO',                d['gasto'])
    set_col('RUBRO',                d['rubro'])
    set_col('MES_IMPUTACION',       d['mes'])
    set_col('ANIO_IMPUTACION',      d['anio'])
    set_col('CODIGO_OPERACION',     d['codigo_operacion'])
    set_col('POSICION',             d['posicion'])
    set_col('DESCRIPCION_GASTO',    d['descripcion_gasto'])
    set_col('DESCRIPCION_RUBRO',    d['descripcion_rubro'])
    set_col('ESTADO',               'OK')
    set_col('HORA_ESTADO',          ahora)
    set_col('NOMBRE_ADJUNTO',       d['nombre_adjunto'])
    set_col('CLAVE_COMPROBANTE',    d['clave'])

    # Marcar celdas que necesitan revisión manual (None = no se pudo extraer)
    def verificar(key, valor):
        if valor is None or valor == '' or valor == 'None':
            fila[COLS[key] - 1] = 'VERIFICAR'
            cols_verificar.append(COLS[key])

    verificar('TIPO_COMPROBANTE', d['tipo'])
    verificar('NUMERO',           d['numero'])
    verificar('FECHA',            d['fecha'])
    verificar('DENOMINACION',     d['denominacion'])
    verificar('CUIT',             d['cuit'])

    if d['kilos_num'] is None or d['kilos_num'] <= 0:
        fila[COLS['KILOS'] - 1] = 'VERIFICAR'
        cols_verificar.append(COLS['KILOS'])

    if d['precio_unitario_num'] is None or d['precio_unitario_num'] <= 0:
        fila[COLS['PRECIO_UNITARIO'] - 1] = 'VERIFICAR'
        cols_verificar.append(COLS['PRECIO_UNITARIO'])

    if d['tipo'] == 'FCA' and d['neto_num'] is None:
        fila[COLS['NETO'] - 1] = 'VERIFICAR'
        cols_verificar.append(COLS['NETO'])

    if d['tipo'] == 'FCA' and d['iva_num'] is None:
        fila[COLS['IVA'] - 1] = 'VERIFICAR'
        cols_verificar.append(COLS['IVA'])

    if d['tipo'] == 'FCA' and d['tasa'] is None:
        fila[COLS['TASA'] - 1] = 'VERIFICAR'
        cols_verificar.append(COLS['TASA'])

    if d['total_num'] is None or d['total_num'] <= 0:
        fila[COLS['TOTAL'] - 1] = 'VERIFICAR'
        cols_verificar.append(COLS['TOTAL'])
    elif d['tipo'] == 'FCA' and d['neto_num'] is not None and d['iva_num'] is not None:
        otros = d.get('otros_tributos_num') or 0
        suma  = d['neto_num'] + d['iva_num'] + otros
        if abs(d['total_num'] - suma) > 1.0:
            for key in ('NETO', 'IVA', 'TOTAL'):
                fila[COLS[key] - 1] = 'VERIFICAR (Suma)'
                if COLS[key] not in cols_verificar:
                    cols_verificar.append(COLS[key])

    if d['gasto'] is None:
        fila[COLS['GASTO'] - 1] = 'VERIFICAR'
        cols_verificar.append(COLS['GASTO'])

    if d['rubro'] is None:
        fila[COLS['RUBRO'] - 1] = 'VERIFICAR'
        cols_verificar.append(COLS['RUBRO'])

    if not d['mes']:
        fila[COLS['MES_IMPUTACION'] - 1] = 'VERIFICAR'
        cols_verificar.append(COLS['MES_IMPUTACION'])

    if not d['anio']:
        fila[COLS['ANIO_IMPUTACION'] - 1] = 'VERIFICAR'
        cols_verificar.append(COLS['ANIO_IMPUTACION'])

    return fila, cols_verificar
