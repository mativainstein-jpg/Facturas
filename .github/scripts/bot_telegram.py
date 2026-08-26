"""
Bot de Telegram (vía GitHub Actions): revisa si llegó un Excel de
proveedores nuevo y actualiza los maestros solo, sin revisión.

Corre en un runner de GitHub Actions (no en la sandbox de Claude), así que
tiene acceso normal a internet. Reutiliza la MISMA lógica de fusión que se
venía aplicando a mano.
"""
import os
import re
import subprocess
import sys
import requests
import openpyxl

TELEGRAM_BOT_TOKEN = os.environ['TELEGRAM_BOT_TOKEN']
AUTHORIZED_ID = os.environ['AUTHORIZED_TELEGRAM_ID']
TARGET_BRANCHES = os.environ.get(
    'TARGET_BRANCHES',
    'main,claude/invoice-app-setup-ss5uwf,claude/bejerman-invoice-reader-do4scq'
).split(',')

STATE_FILE = 'bot_state/last_update_id.txt'
MIN_FILAS_VALIDAS = 100
TG = f'https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}'
GITHUB_STATUS_API = 'https://www.githubstatus.com/api/v2/status.json'


def tg_get(method, **params):
    r = requests.get(f'{TG}/{method}', params=params, timeout=30)
    r.raise_for_status()
    return r.json()


def tg_send(chat_id, text):
    requests.post(f'{TG}/sendMessage', json={'chat_id': chat_id, 'text': text}, timeout=30)


def sh(*args, cwd=None):
    r = subprocess.run(args, cwd=cwd, capture_output=True, text=True)
    if r.returncode != 0:
        raise RuntimeError(f'git {" ".join(args)} -> {r.returncode}: {r.stdout}\n{r.stderr}')
    return r.stdout.strip()


def leer_last_update_id():
    try:
        return int(open(STATE_FILE).read().strip())
    except Exception:
        return 0


def guardar_last_update_id(uid):
    os.makedirs('bot_state', exist_ok=True)
    with open(STATE_FILE, 'w') as f:
        f.write(str(uid))


def descargar_documento(file_id):
    info = tg_get('getFile', file_id=file_id)
    file_path = info['result']['file_path']
    url = f'https://api.telegram.org/file/bot{TELEGRAM_BOT_TOKEN}/{file_path}'
    r = requests.get(url, timeout=60)
    r.raise_for_status()
    return r.content


def chequear_estado_github():
    """
    Si algo falla, esto ayuda a distinguir "GitHub tiene un problema" de
    "el archivo/código tienen un problema real". Corre DENTRO del runner
    de Actions (con internet normal), no depende de la sandbox de Claude.
    """
    try:
        r = requests.get(GITHUB_STATUS_API, timeout=10)
        indicator = r.json().get('status', {}).get('indicator', 'unknown')
        if indicator == 'none':
            return 'GitHub funciona normal (no parece ser una caída de GitHub).'
        return f'⚠️ GitHub reporta un problema activo ahora mismo (nivel: {indicator}).'
    except Exception:
        return 'No pude confirmar el estado de GitHub.'


def parsear_archivo_nuevo(buffer, nombre_archivo):
    # Igual que en las actualizaciones manuales: acepta .xls viejo o .xlsx,
    # busca las columnas por nombre (cuit, nombre, codgasto, codrubro).
    if nombre_archivo.lower().endswith('.xls'):
        import xlrd
        wb = xlrd.open_workbook(file_contents=buffer)
        sheet = wb.sheet_by_index(0)
        headers = [str(h).strip().lower() for h in sheet.row_values(0)]
        idx = {h: i for i, h in enumerate(headers)}
        filas = [sheet.row_values(r) for r in range(1, sheet.nrows)]
    else:
        import io
        wb = openpyxl.load_workbook(io.BytesIO(buffer), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(h).strip().lower() if h else '' for h in rows[0]]
        idx = {h: i for i, h in enumerate(headers)}
        filas = rows[1:]

    for col in ('cuit', 'nombre', 'codgasto', 'codrubro'):
        if col not in idx:
            raise ValueError(
                f'El archivo no tiene la columna "{col}" (busqué cuit, nombre, '
                'codgasto, codrubro). ¿Es el Excel de proveedores correcto?'
            )

    nuevo = {}
    for row in filas:
        cuit_raw = row[idx['cuit']]
        cuit = re.sub(r'\D', '', str(cuit_raw or ''))
        if len(cuit) != 11:
            continue
        nombre = str(row[idx['nombre']] or '').strip()
        gasto_raw = row[idx['codgasto']]
        rubro_raw = row[idx['codrubro']]
        gasto = int(gasto_raw) if gasto_raw not in ('', None) else None
        rubro = int(rubro_raw) if rubro_raw not in ('', None) else None
        nuevo[cuit] = {'nombre': nombre, 'gasto': gasto, 'rubro': rubro}

    if len(nuevo) < MIN_FILAS_VALIDAS:
        raise ValueError(
            f'Solo encontré {len(nuevo)} filas con CUIT válido (esperaba miles). '
            'No actualicé nada — revisá que sea el archivo correcto.'
        )
    return nuevo


def actualizar_nombres(path, nuevo):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    fila_por_cuit = {}
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if row[0].value:
            c = re.sub(r'\D', '', str(row[0].value))
            if c:
                fila_por_cuit[c] = i
    agregados = corregidos = 0
    for cuit, d in nuevo.items():
        if not d['nombre']:
            continue
        if cuit in fila_por_cuit:
            fila = fila_por_cuit[cuit]
            actual = str(ws.cell(fila, 2).value or '').strip()
            if actual != d['nombre']:
                ws.cell(fila, 2).value = d['nombre']
                corregidos += 1
        else:
            ws.append([cuit, d['nombre']])
            agregados += 1
    if agregados + corregidos:
        wb.save(path)
    return agregados, corregidos


def actualizar_proveedores(path, nuevo):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    fila_por_cuit = {}
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if row[0].value:
            c = re.sub(r'\D', '', str(row[0].value))
            if c:
                fila_por_cuit[c] = i
    agregados = corregidos = 0
    for cuit, d in nuevo.items():
        if d['gasto'] is None and d['rubro'] is None:
            continue
        if cuit in fila_por_cuit:
            fila = fila_por_cuit[cuit]
            g_actual = ws.cell(fila, 2).value
            ru_actual = ws.cell(fila, 4).value
            g_actual = int(g_actual) if g_actual not in (None, '') else None
            ru_actual = int(ru_actual) if ru_actual not in (None, '') else None
            if g_actual != d['gasto'] or ru_actual != d['rubro']:
                ws.cell(fila, 2).value = d['gasto']
                ws.cell(fila, 4).value = d['rubro']
                corregidos += 1
        else:
            ws.append([cuit, d['gasto'], None, d['rubro']])
            agregados += 1
    if agregados + corregidos:
        wb.save(path)
    return agregados, corregidos


def actualizar_rama(branch, nuevo):
    sh('git', 'fetch', 'origin', branch)
    sh('git', 'checkout', '-B', branch, f'origin/{branch}')
    an, cn = actualizar_nombres('cuit_nombre.xlsx', nuevo)
    ap, cp = actualizar_proveedores('proveedores.xlsx', nuevo)
    total = an + cn + ap + cp
    if total == 0:
        return {'branch': branch, 'nuevos': 0, 'corregidos': 0}
    sh('git', 'add', 'cuit_nombre.xlsx', 'proveedores.xlsx')
    sh('git', 'commit', '-m',
       f'Actualizar maestros de proveedores vía Telegram '
       f'({an + ap} nuevos, {cn + cp} corregidos)\n\nEnviado por Telegram, aplicado sin revisión.')
    sh('git', 'push', 'origin', branch)
    return {'branch': branch, 'nuevos': an + ap, 'corregidos': cn + cp}


def procesar_mensaje(chat_id, doc):
    tg_send(chat_id, f'📥 Recibido: {doc.get("file_name", "archivo")}. Lo estoy revisando...')

    try:
        buffer = descargar_documento(doc['file_id'])
        nuevo_map = parsear_archivo_nuevo(buffer, doc.get('file_name', 'archivo.xlsx'))
    except Exception as e:
        estado_gh = chequear_estado_github()
        tg_send(chat_id, f'❌ No pude leer el archivo: {e}\n\n{estado_gh}')
        return

    tg_send(chat_id, f'⏳ En proceso: actualizando {len(nuevo_map)} proveedores en cada rama...')

    try:
        resultados = [actualizar_rama(b.strip(), nuevo_map) for b in TARGET_BRANCHES]
    except Exception as e:
        estado_gh = chequear_estado_github()
        tg_send(chat_id, f'❌ No pude actualizar: {e}\n\n{estado_gh}')
        return

    r0 = resultados[0]
    if r0['nuevos'] + r0['corregidos'] == 0:
        resumen = f'✅ Hecho. Revisé {len(nuevo_map)} proveedores, no había nada nuevo para actualizar.'
    else:
        resumen = (
            f'✅ Hecho (leí {len(nuevo_map)} filas):\n'
            f'• {r0["nuevos"]} nuevos\n'
            f'• {r0["corregidos"]} corregidos\n'
            f'Ramas: {", ".join(r["branch"] for r in resultados)}'
        )
    tg_send(chat_id, resumen)


def main():
    sh('git', 'config', 'user.email', 'bot@naiman.local')
    sh('git', 'config', 'user.name', 'Bot Proveedores')

    last_id = leer_last_update_id()
    updates = tg_get('getUpdates', offset=last_id + 1, timeout=0)['result']

    if not updates:
        print('Sin mensajes nuevos.')
        return

    nuevo_last_id = last_id
    for upd in updates:
        nuevo_last_id = max(nuevo_last_id, upd['update_id'])
        msg = upd.get('message')
        if not msg:
            continue
        from_id = msg.get('from', {}).get('id')
        if str(from_id) != str(AUTHORIZED_ID):
            continue  # ignorar en silencio a cualquiera que no sea la persona autorizada
        chat_id = msg['chat']['id']
        doc = msg.get('document')
        if not doc:
            tg_send(chat_id, 'Mandame el Excel de proveedores como archivo adjunto.')
            continue
        procesar_mensaje(chat_id, doc)

    # Volver a main y guardar el último update_id procesado
    sh('git', 'fetch', 'origin', 'main')
    sh('git', 'checkout', '-B', 'main', 'origin/main')
    guardar_last_update_id(nuevo_last_id)
    sh('git', 'add', STATE_FILE)
    r = subprocess.run(['git', 'diff', '--cached', '--quiet'])
    if r.returncode != 0:  # hay cambios para commitear
        sh('git', 'commit', '-m', f'Bot: marcar update_id {nuevo_last_id} como procesado')
        sh('git', 'push', 'origin', 'main')


if __name__ == '__main__':
    main()
